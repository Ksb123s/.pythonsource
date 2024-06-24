# 엑셀 저장 모듈
from openpyxl import Workbook
from random import randint
from openpyxl.drawing.image import Image 
from openpyxl import load_workbook
from openpyxl.chart import BarChart,Reference,LineChart
from openpyxl.styles import Font,Border, PatternFill, Alignment
from openpyxl.styles.borders import Side
from datetime import datetime

def write_excel_template(filename, sheetname, listdata):
    wb = Workbook()
    ws = wb.active

    ws.title = sheetname
    # 너비 조정
    ws.column_dimensions["A"].width = 100
    ws.column_dimensions["B"].width = 30
    for data in listdata:
        ws.append(data)
    base_url = "C:\\source\\pythonsource\\crawl\\file\\"


    wb.save( base_url + filename)
    wb.close()


if __name__ == "__main__":
    listdata = [["이름", "나이"], ["홍길동", "25"], ["성춘향", 22]]
    write_excel_template("temp.xlsx", "test", listdata)